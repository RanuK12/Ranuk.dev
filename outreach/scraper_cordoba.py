"""
Ranuk IT — Lead generator para empresas de Córdoba, AR.
Uso:
    python scraper_cordoba.py --rubro agroindustria --max 80 --out leads_agro.xlsx

Requisitos:
    pip install httpx beautifulsoup4 pandas openpyxl

API keys necesarias (cargalas en config/.env):
    GOOGLE_MAPS_API_KEY   → https://console.cloud.google.com (habilitar Places API New)
    PAGESPEED_API_KEY     → misma consola (habilitar PageSpeed Insights API)

Costo aproximado por corrida de 80 empresas:
    Places New: ~USD 1.40  /  PageSpeed: gratis hasta 25k/día

Cumplimiento legal (AR):
    Ley 25.326 permite prospección B2B a emails corporativos con opt-out claro
    en cada email. NO usar este script para prospección a emails personales.
"""

import argparse
import asyncio
import json
import os
import re
from dataclasses import dataclass, field

import httpx
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ==========================================================================
# CONFIG
# ==========================================================================
GOOGLE_MAPS_API_KEY = os.getenv("GOOGLE_MAPS_API_KEY", "PEGAR_API_KEY_AQUI")
PAGESPEED_API_KEY = os.getenv("PAGESPEED_API_KEY", "PEGAR_API_KEY_AQUI")
USER_AGENT = "Mozilla/5.0 (compatible; RanukBot/1.0; +https://ranuk.dev)"
TIMEOUT = 15

# Bounding box aproximado de la provincia de Córdoba, AR
CORDOBA_BBOX = {
    "low":  {"latitude": -34.9, "longitude": -65.8},
    "high": {"latitude": -29.5, "longitude": -61.8},
}

# Queries por rubro — podés sumar más combinaciones
RUBROS = {
    "agroindustria": [
        "agroindustria cordoba argentina",
        "cerealera cordoba argentina",
        "acopio cereales cordoba",
        "feedlot cordoba argentina",
        "exportador granos cordoba",
    ],
    "manufactura": [
        "metalurgica cordoba argentina",
        "autopartes cordoba argentina",
        "fabrica plastico cordoba",
        "industria alimenticia cordoba",
    ],
    "clinicas": [
        "clinica privada cordoba argentina",
        "centro medico cordoba",
        "laboratorio analisis cordoba",
        "sanatorio cordoba argentina",
    ],
    "logistica": [
        "transporte cargas cordoba argentina",
        "logistica cordoba argentina",
        "distribuidora cordoba argentina",
        "empresa courier cordoba",
    ],
    "retail": [
        "cadena supermercados cordoba",
        "cadena ferreterias cordoba",
        "cadena farmacias cordoba",
    ],
}

# Señales de tech debt detectables en HTML/headers
LEGACY_SIGNALS = {
    "jquery-1.":                 "jQuery v1 (legacy 10+ años)",
    "jquery-2.":                 "jQuery v2 (deprecado)",
    "bootstrap-3":               "Bootstrap 3 (deprecado 2019)",
    "wp-content/plugins/revslider": "Revolution Slider (CVEs conocidos)",
    "php/5.":                    "PHP 5.x (EOL desde 2018)",
    "x-powered-by: asp.net":     "ASP.NET clásico sin versionar",
    ".aspx?":                    "WebForms ASP.NET (stack legacy)",
    "frontpage":                 "Microsoft FrontPage (abandonado 2006)",
    "flash":                     "Adobe Flash (EOL 2020)",
    "joomla!":                   "Joomla (muchas instalaciones desactualizadas)",
    "powered by drupal 7":       "Drupal 7 (EOL enero 2025)",
}

# ==========================================================================
# MODELO
# ==========================================================================
@dataclass
class Lead:
    empresa: str = ""
    rubro: str = ""
    sitio: str = ""
    email: str = ""
    telefono: str = ""
    direccion: str = ""
    descripcion: str = ""
    falencia_tecnologica: str = ""
    stack_detectado: str = ""
    pagespeed_mobile: int = -1
    tiene_https: bool = False
    score_oportunidad: int = 0
    fuente: str = "google_places"
    notas: str = ""


# ==========================================================================
# 1. DESCUBRIR EMPRESAS VÍA GOOGLE PLACES
# ==========================================================================
async def discover_companies(client: httpx.AsyncClient, query: str, max_results: int = 60):
    """Usa Places API (New) — Text Search."""
    if GOOGLE_MAPS_API_KEY.startswith("PEGAR"):
        raise RuntimeError("Falta GOOGLE_MAPS_API_KEY. Configurá la env var.")

    url = "https://places.googleapis.com/v1/places:searchText"
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": GOOGLE_MAPS_API_KEY,
        "X-Goog-FieldMask": (
            "places.displayName,places.websiteUri,places.formattedAddress,"
            "places.nationalPhoneNumber,places.types,places.businessStatus,"
            "places.rating,places.userRatingCount"
        ),
    }

    results = []
    page_token = None
    for _ in range(4):  # máximo 4 páginas × 20 = 80 resultados
        body = {
            "textQuery": query,
            "pageSize": 20,
            "locationBias": {"rectangle": CORDOBA_BBOX},
        }
        if page_token:
            body["pageToken"] = page_token

        r = await client.post(url, headers=headers, json=body)
        if r.status_code != 200:
            print(f"  ⚠ Places API {r.status_code}: {r.text[:200]}")
            break

        data = r.json()
        batch = data.get("places", [])
        results.extend(batch)
        page_token = data.get("nextPageToken")

        if not page_token or len(results) >= max_results:
            break
        await asyncio.sleep(2)

    return results[:max_results]


# ==========================================================================
# 2. SCRAPE SITIO + EMAIL
# ==========================================================================
EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
EMAIL_BLACKLIST = (
    "wixpress", "sentry", "example.com", "noreply", "godaddy",
    "privacy", "tu-email", "your-email", ".png", ".jpg", ".svg", ".gif",
)
CONTACT_PATHS = ["/", "/contacto", "/contactanos", "/contact", "/nosotros", "/about", "/institucional"]


async def scrape_site(client: httpx.AsyncClient, site_url: str):
    """Busca emails y señales tech debt. Devuelve (emails, html, headers)."""
    base = site_url.rstrip("/")
    emails = set()
    html_blob = ""
    headers_blob = {}

    for path in CONTACT_PATHS:
        try:
            r = await client.get(
                base + path,
                headers={"User-Agent": USER_AGENT},
                follow_redirects=True,
            )
            if r.status_code != 200:
                continue
            headers_blob.update({k.lower(): v for k, v in r.headers.items()})
            html_blob += r.text[:400_000]  # cap por URL para no explotar memoria
            for m in EMAIL_RE.findall(r.text):
                m = m.lower()
                if any(bad in m for bad in EMAIL_BLACKLIST):
                    continue
                emails.add(m)
            # Detectar mailto:
            for a in BeautifulSoup(r.text, "html.parser").select("a[href^=mailto]"):
                href = a.get("href", "").replace("mailto:", "").split("?")[0].strip().lower()
                if href and "@" in href and not any(b in href for b in EMAIL_BLACKLIST):
                    emails.add(href)
        except Exception:
            continue

    # Priorizar emails de dominio propio
    domain = re.sub(r"^https?://(www\.)?", "", site_url).split("/")[0]
    sorted_emails = sorted(
        emails,
        key=lambda e: (0 if domain in e else 1, len(e)),
    )
    return sorted_emails[:3], html_blob, headers_blob


# ==========================================================================
# 3. DETECTAR FALENCIAS TÉCNICAS
# ==========================================================================
def detect_tech_debt(html: str, headers: dict) -> tuple[list[str], list[str]]:
    blob = html.lower() + " " + " ".join(f"{k}:{v}" for k, v in headers.items()).lower()
    falencias: list[str] = []
    stack: list[str] = []

    for signal, human in LEGACY_SIGNALS.items():
        if signal.lower() in blob:
            falencias.append(human)

    # Stack detection
    if "wp-content" in blob:
        stack.append("WordPress")
    if "shopify" in blob or "cdn.shopify" in blob:
        stack.append("Shopify")
    if "wix.com" in blob or "wixstatic" in blob:
        stack.append("Wix")
    if "squarespace" in blob:
        stack.append("Squarespace")
    if "react" in blob and "__next" in blob:
        stack.append("Next.js")
    elif "react" in blob:
        stack.append("React")
    if "vue" in blob:
        stack.append("Vue")

    # Security headers
    if "strict-transport-security" not in headers:
        falencias.append("Sin HSTS (seguridad HTTP incompleta)")
    if "content-security-policy" not in headers:
        falencias.append("Sin Content-Security-Policy")

    # Portal / área de clientes ausente en stack obsoleto
    if "wordpress" in (s.lower() for s in stack) and not re.search(
        r"\b(login|portal|mi[- ]cuenta|clientes|acceso|intranet)\b", blob
    ):
        falencias.append("Sin portal de clientes / área privada detectable")

    # Pagos online ausentes en retail
    if any(w in blob for w in ["tienda", "productos", "carrito"]) and not any(
        p in blob for p in ["mercadopago", "mercado pago", "stripe", "paypal", "checkout"]
    ):
        falencias.append("E-commerce sin pasarela de pagos moderna")

    return falencias[:3], stack


# ==========================================================================
# 4. PAGESPEED MOBILE
# ==========================================================================
async def pagespeed_mobile(client: httpx.AsyncClient, url: str) -> int:
    if PAGESPEED_API_KEY.startswith("PEGAR"):
        return -1
    try:
        r = await client.get(
            "https://www.googleapis.com/pagespeedonline/v5/runPagespeed",
            params={
                "url": url,
                "strategy": "mobile",
                "key": PAGESPEED_API_KEY,
                "category": "performance",
            },
            timeout=45,
        )
        data = r.json()
        return int(data["lighthouseResult"]["categories"]["performance"]["score"] * 100)
    except Exception:
        return -1


# ==========================================================================
# 5. SCORING
# ==========================================================================
def opportunity_score(lead: Lead) -> int:
    """0-100 heurística para priorizar leads en el outreach."""
    score = 25  # base

    if lead.email:
        score += 20
    if lead.falencia_tecnologica and "moderno" not in lead.falencia_tecnologica.lower():
        score += 25
    if 0 <= lead.pagespeed_mobile < 50:
        score += 15
    elif 0 <= lead.pagespeed_mobile < 70:
        score += 8
    if not lead.tiene_https:
        score += 10
    if lead.telefono:
        score += 5

    return min(score, 100)


# ==========================================================================
# 6. PIPELINE
# ==========================================================================
async def build_leads(rubro: str, max_empresas: int) -> list[Lead]:
    queries = RUBROS.get(rubro, [rubro])
    per_query = max(10, max_empresas // len(queries))

    async with httpx.AsyncClient(timeout=TIMEOUT, follow_redirects=True) as client:
        # 6.1 discovery
        all_places = []
        seen_names = set()
        for q in queries:
            print(f"🔍 Buscando: {q}")
            for p in await discover_companies(client, q, max_results=per_query):
                name = p.get("displayName", {}).get("text", "")
                if name and name not in seen_names:
                    seen_names.add(name)
                    all_places.append(p)

        print(f"→ {len(all_places)} empresas únicas encontradas. Analizando sitios...\n")

        # 6.2 enriquecimiento
        leads: list[Lead] = []
        for idx, p in enumerate(all_places, 1):
            name = p.get("displayName", {}).get("text", "—")
            site = (p.get("websiteUri") or "").strip()

            lead = Lead(
                empresa=name,
                rubro=rubro,
                sitio=site,
                telefono=p.get("nationalPhoneNumber", ""),
                direccion=p.get("formattedAddress", ""),
                descripcion=", ".join(p.get("types", [])[:3]),
            )

            if not site:
                lead.notas = "Sin sitio web — contactar por teléfono o buscar en LinkedIn"
                lead.score_oportunidad = opportunity_score(lead)
                leads.append(lead)
                print(f"[{idx:03d}] {name[:50]:50s}  ⚠ sin sitio")
                continue

            try:
                emails, html, headers = await scrape_site(client, site)
                falencias, stack = detect_tech_debt(html, headers)
                ps = await pagespeed_mobile(client, site)

                lead.email = emails[0] if emails else ""
                lead.falencia_tecnologica = " · ".join(falencias) if falencias else "Stack razonable — priorizar otra"
                lead.stack_detectado = ", ".join(stack)
                lead.pagespeed_mobile = ps
                lead.tiene_https = site.startswith("https")
                lead.score_oportunidad = opportunity_score(lead)

                leads.append(lead)
                print(f"[{idx:03d}] {name[:50]:50s}  score={lead.score_oportunidad:3d}  {'✓' if emails else '✗'}email  ps={ps}")
            except Exception as e:
                lead.notas = f"Error scraping: {e}"
                leads.append(lead)
                print(f"[{idx:03d}] {name[:50]:50s}  ✗ error: {e}")

    return sorted(leads, key=lambda l: -l.score_oportunidad)


# ==========================================================================
# 7. EXPORT XLSX
# ==========================================================================
def export_xlsx(leads: list[Lead], path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads Córdoba"

    headers = [
        "Empresa", "Rubro", "Sitio", "Email", "Teléfono", "Dirección",
        "Descripción", "Falencia técnica", "Stack detectado",
        "PageSpeed móvil", "HTTPS", "Score", "Fuente", "Notas",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="0A192F")
    for c in ws[1]:
        c.font = Font(bold=True, color="D4A574", size=11)
        c.fill = header_fill
        c.alignment = Alignment(vertical="center")

    for l in leads:
        ws.append([
            l.empresa, l.rubro, l.sitio, l.email, l.telefono, l.direccion,
            l.descripcion, l.falencia_tecnologica, l.stack_detectado,
            l.pagespeed_mobile if l.pagespeed_mobile >= 0 else "—",
            "Sí" if l.tiene_https else "No",
            l.score_oportunidad, l.fuente, l.notas,
        ])

    # Colorear filas por score
    gold_soft = PatternFill("solid", fgColor="FFF7E6")
    green_soft = PatternFill("solid", fgColor="E8F5E9")
    for row_idx, l in enumerate(leads, start=2):
        if l.score_oportunidad >= 75:
            for c in ws[row_idx]:
                c.fill = green_soft
        elif l.score_oportunidad >= 60:
            for c in ws[row_idx]:
                c.fill = gold_soft

    # Ancho de columnas
    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=10) + 2
        ws.column_dimensions[col[0].column_letter].width = min(width, 55)

    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"\n📊 {len(leads)} leads exportados → {path}")
    print(f"   → Alta prioridad (score ≥75): {sum(1 for l in leads if l.score_oportunidad >= 75)}")
    print(f"   → Media  (60-74):              {sum(1 for l in leads if 60 <= l.score_oportunidad < 75)}")
    print(f"   → Baja   (<60):                {sum(1 for l in leads if l.score_oportunidad < 60)}")


# ==========================================================================
# MAIN
# ==========================================================================
def main():
    ap = argparse.ArgumentParser(description="Lead generator para Ranuk IT — empresas de Córdoba AR")
    ap.add_argument("--rubro", required=True, choices=list(RUBROS.keys()),
                    help="Rubro a prospectar")
    ap.add_argument("--max", type=int, default=60,
                    help="Máximo de empresas a procesar (default: 60)")
    ap.add_argument("--out", default=None,
                    help="Archivo de salida .xlsx")
    args = ap.parse_args()

    out = args.out or f"leads_{args.rubro}.xlsx"
    leads = asyncio.run(build_leads(args.rubro, args.max))
    export_xlsx(leads, out)


if __name__ == "__main__":
    main()
